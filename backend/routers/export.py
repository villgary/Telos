"""
Report Export API — Excel export for review reports and compliance runs.
"""

import io
from datetime import datetime, timezone
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from backend.database import get_db
from backend import models, auth

router = APIRouter(prefix="/api/v1/export", tags=["报表导出"])


def _style_header(ws, row=1):
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F3864")
    thin = Side(style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[row]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border


def _auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                val = str(cell.value) if cell.value is not None else ""
                max_len = max(max_len, len(val))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)


# ── Review Report Export ────────────────────────────────────────────────────────

@router.get("/review-report/{report_id}")
async def export_review_report(
    report_id: int,
    db: Session = Depends(get_db),
    user: models.User = Depends(auth.get_current_user),
):
    """
    Export a review report as Excel.
    Sheet1: summary, Sheet2: account details.
    """
    report = db.query(models.ReviewReport).filter(
        models.ReviewReport.id == report_id
    ).first()
    if not report:
        raise HTTPException(status_code=404, detail="报告不存在")

    summary = report.content_summary or {}
    s = summary.get("summary", {})

    wb = Workbook()

    # ── Sheet1: Summary ──────────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "摘要"
    ws1.append(["审查报告导出"])
    ws1.append([])

    meta = [
        ("审查计划", report.schedule.name if report.schedule else "-"),
        ("周期", report.period),
        ("周期起始", report.period_start.strftime("%Y-%m-%d") if report.period_start else "-"),
        ("周期结束", report.period_end.strftime("%Y-%m-%d") if report.period_end else "-"),
        ("状态", report.status),
        ("生成时间", report.created_at.strftime("%Y-%m-%d %H:%M") if report.created_at else "-"),
        ("审核人", report.reviewer.username if report.reviewer else "-"),
        ("审核时间", report.reviewed_at.strftime("%Y-%m-%d %H:%M") if report.reviewed_at else "-"),
        ("备注", report.notes or "-"),
    ]
    for label, value in meta:
        ws1.append([label, value])

    ws1.append([])
    stats = [
        ("总账号数", s.get("total_accounts", 0)),
        ("特权账号数", s.get("privileged_count", 0)),
        ("休眠账号数", s.get("dormant_count", 0)),
        ("离机账号数", s.get("departed_count", 0)),
        ("高风险资产数", s.get("high_risk_count", 0)),
    ]
    ws1.append(["统计项", "数量"])
    for label, value in stats:
        ws1.append([label, value])

    ws1.column_dimensions["A"].width = 18
    ws1.column_dimensions["B"].width = 18

    # ── Sheet2: Dormant accounts ────────────────────────────────────────────────
    dormant = summary.get("dormant_accounts", [])
    if dormant:
        ws2 = wb.create_sheet("休眠账号")
        ws2.append(["用户名", "UID/SID", "资产", "IP", "主机名", "上次登录", "状态"])
        for a in dormant:
            ws2.append([
                a.get("username", ""),
                a.get("uid", ""),
                a.get("asset_code", ""),
                a.get("ip", ""),
                a.get("hostname", ""),
                a.get("last_login", "")[:10] if a.get("last_login") else "",
                a.get("status", ""),
            ])
        _style_header(ws2)
        _auto_width(ws2)

    # ── Sheet3: Departed accounts ───────────────────────────────────────────────
    departed = summary.get("departed_accounts", [])
    if departed:
        ws3 = wb.create_sheet("离机账号")
        ws3.append(["用户名", "UID/SID", "资产", "IP", "主机名", "上次登录", "状态"])
        for a in departed:
            ws3.append([
                a.get("username", ""),
                a.get("uid", ""),
                a.get("asset_code", ""),
                a.get("ip", ""),
                a.get("hostname", ""),
                a.get("last_login", "")[:10] if a.get("last_login") else "",
                a.get("status", ""),
            ])
        _style_header(ws3)
        _auto_width(ws3)

    # ── Sheet4: High-risk assets ───────────────────────────────────────────────
    high_risk = summary.get("high_risk_assets", [])
    if high_risk:
        ws4 = wb.create_sheet("高风险资产")
        ws4.append(["资产编号", "IP", "风险分", "风险等级", "风险因子"])
        for a in high_risk:
            factors = ", ".join(a.get("risk_factors", [])[:3])
            ws4.append([
                a.get("asset_code", ""),
                a.get("ip", ""),
                a.get("risk_score", ""),
                a.get("risk_level", ""),
                factors,
            ])
        _style_header(ws4)
        _auto_width(ws4)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"审查报告_{report.schedule.name if report.schedule else report_id}_{datetime.now(timezone.utc).strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename*=UTF-8\'\'{filename}'},
    )


# ── Compliance Run Export ───────────────────────────────────────────────────────

@router.get("/compliance-run/{run_id}")
async def export_compliance_run(
    run_id: int,
    db: Session = Depends(get_db),
    user: models.User = Depends(auth.get_current_user),
):
    """
    Export a compliance run as Excel.
    Sheet1: framework summary, Sheet2: detailed results.
    """
    run = db.query(models.ComplianceRun).filter(
        models.ComplianceRun.id == run_id
    ).first()
    if not run:
        raise HTTPException(status_code=404, detail="合规运行不存在")

    results = db.query(models.ComplianceResult).filter(
        models.ComplianceResult.run_id == run_id
    ).all()

    wb = Workbook()

    # ── Sheet1: Summary ──────────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "框架摘要"
    ws1.append([f"合规报告 — {run.framework_slug}"])
    ws1.append([])
    ws1.append(["运行ID", run.id])
    ws1.append(["框架", run.framework_slug])
    ws1.append(["运行时间", run.started_at.strftime("%Y-%m-%d %H:%M") if run.started_at else "-"])
    ws1.append(["完成时间", run.completed_at.strftime("%Y-%m-%d %H:%M") if run.completed_at else "-"])
    ws1.append(["状态", run.status])
    passed = sum(1 for r in results if r.result == "pass")
    failed = sum(1 for r in results if r.result == "fail")
    ws1.append(["通过数", passed])
    ws1.append(["失败数", failed])
    ws1.append(["通过率", f"{passed}/{passed+failed} ({round(passed/(passed+failed)*100) if passed+failed else 0}%)"])
    ws1.column_dimensions["A"].width = 18
    ws1.column_dimensions["B"].width = 30

    # ── Sheet2: Results ───────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("检查结果")
    ws2.append(["资产编号", "IP", "检查项", "结果", "控制项ID", "描述", "证据"])
    for r in results:
        asset = db.query(models.Asset).filter(models.Asset.id == r.asset_id).first() if r.asset_id else None
        evidence_str = ""
        if isinstance(r.evidence, dict):
            evidence_str = str(r.evidence)[:200]
        elif isinstance(r.evidence, list):
            evidence_str = "; ".join(str(e) for e in r.evidence[:5])
        elif r.evidence:
            evidence_str = str(r.evidence)[:200]
        ws2.append([
            asset.asset_code if asset else "-",
            asset.ip if asset else "-",
            r.check_title or "-",
            r.result,
            r.control_id or "-",
            r.description or "-",
            evidence_str,
        ])
    _style_header(ws2)
    for col in ws2.columns:
        col_letter = col[0].column_letter
        ws2.column_dimensions[col_letter].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"合规报告_{run.framework_slug}_{datetime.now(timezone.utc).strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename*=UTF-8\'\'{filename}'},
    )
